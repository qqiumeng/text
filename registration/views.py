from django.shortcuts import render,redirect,HttpResponse
from hospital.models import Patient
from doctor.models import  Doctor
from django.http import JsonResponse
from registration.models import Departments
from django import views
from django.core.paginator import Paginator #导入分页器
from django.db.models import Q
import datetime
import openpyxl
# 挂号管理系统


class guahao(views.View):
    def get(self,request):
        a = Patient.objects.all()  # 获取病人所有数据
        paginator = Paginator(a, 5)
        page = int(request.GET.get('page', 1))
        # 分页功能
        if paginator.num_pages > 5:
            if page - 2 < 1:
                page_ran = range(1, 6)
            elif page + 2 > paginator.num_pages:
                page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
            else:
                page_ran = range(page - 2, page + 3)
            page_range = page_ran
        else:
            page_range = paginator.page_range

        contacts = paginator.get_page(page)
        return render(request, 'registrations/index.html',
                      {'contacts': contacts, 'paginator': paginator, 'page_range': page_range,'a':a})
    def post(self,request):


        blh = request.POST.get('blh')
        zzys = request.POST.get('zzys')
        ks = request.POST.get('ks')
        ghsj = request.POST.get('ghsj')
        ghsj2 = request.POST.get('ghsj2')
        if len(blh)>0:
            a = Patient.objects.filter(id=blh)
            paginator = Paginator(a, 5)
            page = int(request.GET.get('page', 1))
            # 分页功能
            if paginator.num_pages > 5:
                if page - 2 < 1:
                    page_ran = range(1, 6)
                elif page + 2 > paginator.num_pages:
                    page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
                else:
                    page_ran = range(page - 2, page + 3)
                page_range = page_ran
            else:
                page_range = paginator.page_range

            contacts = paginator.get_page(page)
            return render(request, 'registrations/index.html',
                          {'contacts': contacts, 'paginator': paginator, 'page_range': page_range})
        elif len(zzys)>0:
            d = Doctor.objects.filter(doctor_name__icontains=zzys)
            if d:
                a = Patient.objects.filter(doctor_id=d[0])
                paginator = Paginator(a, 5)
                page = int(request.GET.get('page', 1))
                # 分页功能
                if paginator.num_pages > 5:
                    if page - 2 < 1:
                        page_ran = range(1, 6)
                    elif page + 2 > paginator.num_pages:
                        page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
                    else:
                        page_ran = range(page - 2, page + 3)
                    page_range = page_ran
                else:
                    page_range = paginator.page_range

                contacts = paginator.get_page(page)
                return render(request, 'registrations/index.html',
                              {'contacts': contacts, 'paginator': paginator, 'page_range': page_range})
            else:
                a = Doctor.objects.filter(doctor_name__icontains=zzys)
                paginator = Paginator(a, 5)
                page = int(request.GET.get('page', 1))
                # 分页功能
                if paginator.num_pages > 5:
                    if page - 2 < 1:
                        page_ran = range(1, 6)
                    elif page + 2 > paginator.num_pages:
                        page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
                    else:
                        page_ran = range(page - 2, page + 3)
                    page_range = page_ran
                else:
                    page_range = paginator.page_range

                contacts = paginator.get_page(page)
                return render(request, 'registrations/index.html',
                              {'contacts': contacts, 'paginator': paginator, 'page_range': page_range})


        elif len(ks)>0:
            d = Departments.objects.filter(department_name__icontains=ks)
            if d:
                a = Patient.objects.filter(doctor_id=d[0])
                paginator = Paginator(a, 5)
                page = int(request.GET.get('page', 1))
                # 分页功能
                if paginator.num_pages > 5:
                    if page - 2 < 1:
                        page_ran = range(1, 6)
                    elif page + 2 > paginator.num_pages:
                        page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
                    else:
                        page_ran = range(page - 2, page + 3)
                    page_range = page_ran
                else:
                    page_range = paginator.page_range

                contacts = paginator.get_page(page)
                return render(request, 'registrations/index.html',
                              {'contacts': contacts, 'paginator': paginator, 'page_range': page_range})
            else:
                a = Departments.objects.filter(department_name__icontains=ks)
                paginator = Paginator(a, 5)
                page = int(request.GET.get('page', 1))
                # 分页功能
                if paginator.num_pages > 5:
                    if page - 2 < 1:
                        page_ran = range(1, 6)
                    elif page + 2 > paginator.num_pages:
                        page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
                    else:
                        page_ran = range(page - 2, page + 3)
                    page_range = page_ran
                else:
                    page_range = paginator.page_range

                contacts = paginator.get_page(page)
                return render(request, 'registrations/index.html',
                              {'contacts': contacts, 'paginator': paginator, 'page_range': page_range})

        elif len(ghsj)>0 and len(ghsj2)>0:

            year = ghsj[0:4]
            month = ghsj[5:7]
            day = ghsj[8:10]
            year2 = ghsj2[0:4]
            month2 = ghsj2[5:7]
            day2 = ghsj2[8:10]

            g1=datetime.date(int(year),int(month),int(day))
            g2=datetime.date(int(year2),int(month2),int(day2))
            a = Patient.objects.filter(g_time__range=(g1, g2))
            paginator = Paginator(a, 5)
            page = int(request.GET.get('page', 1))
            # 分页功能
            if paginator.num_pages > 5:
                if page - 2 < 1:
                    page_ran = range(1, 6)
                elif page + 2 > paginator.num_pages:
                    page_ran = range(paginator.num_pages - 4, paginator.num_pages + 1)
                else:
                    page_ran = range(page - 2, page + 3)
                page_range = page_ran
            else:
                page_range = paginator.page_range

            contacts = paginator.get_page(page)
            return render(request, 'registrations/index.html',
                            {'contacts': contacts, 'paginator': paginator, 'page_range': page_range})
        else:
            return redirect('registrations:guahao')



def get_doctor(request):
    d_name_list = []
    data = {}
    data['status'] = 'SUCCESS'
    office = request.POST.get('d_val')
    print('office在这热', office)
    o_id = Departments.objects.get(department_name=office).id
    print('o_id在这儿', o_id)
    doctors = Doctor.objects.filter(d_office_id=o_id).values_list('doctor_name', flat=True)
    for d_name in doctors:
        d_name_list.append(d_name)
    data['doctors'] = d_name_list
    return JsonResponse(data)



# 挂号更改
class guahaogenggai(views.View):
    def get(self,request):
        content = {}
        content['offices'] = Departments.objects.all()
        return render(request, 'registrations/add.html', content)
    def post(self,request):

        uname = request.POST.get('uname')
        uidn = request.POST.get('uidn')
        umoney = request.POST.get('umoney')
        ussn = request.POST.get('ussn')
        utel = request.POST.get('utel')
        usf = request.POST.get('usf')
        usex = request.POST.get('usex')
        uage = request.POST.get('uage')
        uoc = request.POST.get('uoc')
        ucfz = request.POST.get('ucfz')
        # offic = request.POST.get('offic')
        doctor = request.POST.get('doctor')
        ub = request.POST.get('ub')
        d_id = Doctor.objects.get(doctor_name=doctor).id
        Patient.objects.create(name=uname,id_num=uidn,money=float(umoney),
                                   social_security_num=ussn,tel=utel,self_pay=usf,
                                   sex=usex,age=uage,occupation=uoc,visit=ucfz,remark=ub,
                                   doctor_id_id=d_id)
        return redirect('registrations:guahao')

# 挂号详情

def gaohaoxiangqing(request,u_id):


    u = Patient.objects.filter(id=u_id)
    return render(request, 'registrations/look.html',{'u':u})
 
# 挂号编辑
class guahaoedit(views.View):
    def get(self,request,p_id):
        p = Patient.objects.filter(id=p_id)
        content = {'p':p}
        content['offices'] = Departments.objects.all()

        return render(request, 'registrations/edit.html',content)
    def post(self,request,p_id):
        pname = request.POST.get('pname') #获取输入名字
        pidnum = request.POST.get('pidnum') #获取输入的身份证
        pn = request.POST.get('pn') #获取输入的挂号费
        pssn = request.POST.get('pssn') #获取输入的社保号
        ptel = request.POST.get('ptel')#获取输入的电话号
        pzf = request.POST.get('pzf') #获取单选框的选项 是否自费
        psex = request.POST.get('psex') #获取单选框内的选项 性别男/女
        page = request.POST.get('page') #获取输入的年龄
        pooc = request.POST.get('pooc')#获取输入的职业
        pcfz = request.POST.get('pcfz') #获取单选框内的选项 初诊/复诊
        offic = request.POST.get('offic') #获取所挂科室(医生表)
        doctor = request.POST.get('doctor') #获取指定医生(医生表)
        pbz = request.POST.get('pbz') #获取输入的备注
        p = Patient.objects.get(id=p_id)
        p.name=pname
        p.id_num=pidnum
        p.g_money=pn
        p.social_security_num=pssn
        p.tel=ptel
        p.self_pay=pzf
        p.sex = psex
        p.age = page
        p.occupation=pooc
        p.visit = pcfz
        p.remark=pbz
        d = Doctor.objects.filter(doctor_name=doctor)
        d2 = Departments.objects.filter(department_name=offic)
        if d and d2:
            p.doctor_id = d[0]
            p.save()
        return redirect('registrations:guahao')




def export(request):
    id_list = []
    ids = request.POST.get('ss')
    print(ids)
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in ids:
        if i != ',':
            id_list.append(i)
        else:
            pass
    for i in id_list:
        d = Patient.objects.filter(id=int(i))
        ws['A1'].value = '门诊编号'
        ws['B1'].value = '主治医生'
        ws['C1'].value = '挂号时间'
        ws['D1'].value = '挂号科室'
        ws['E1'].value = '状态'
        row = 2
        for obj in d:
            ws['A' + str(row)].value = obj.id
            ws['B' + str(row)].value = obj.doctor_id.doctor_name
            ws['C' + str(row)].value = obj.g_time
            ws['D' + str(row)].value = obj.doctor_id.d_office.department_name
            ws['E' + str(row)].value = obj.state

            row += 1
            wb.save(f'{obj.name}.xlsx')
            f = open(f'{obj.name}.xlsx', 'rb')
            p = HttpResponse()
            p.content = f
            p['Content-Type'] = 'application/octet-stream'
            p['Content-Disposition'] = 'attachment;filename="%s.xlsx"'% obj.name
    return render(request, 'registrations/index.html')

def exit(request,k_id):
    Patient.objects.filter(id=k_id).update(state='已退号')

    return redirect('registrations:guahao')

def exit2(request):
    id_list = []
    ids = request.POST.get('ss')
    print('ids=', ids)
    for i in ids:
        if i != ',':
            id_list.append(i)
        else:
            pass
    for i in id_list:
        Patient.objects.filter(id=int(i)).update(state='已退号')
    return redirect('registrations:guahao')
