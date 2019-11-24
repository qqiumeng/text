from django.shortcuts import render,redirect,HttpResponse
from django import views
from medicine.forms import *
from django.core.paginator import Paginator
import openpyxl
from django.http import JsonResponse
# Create your views here.


# 增加药品
class add(views.View):
    def get(self,request):
        return render(request, 'medicine/add.html')
    def post(self,request):
        # 获取编号
        drug_num=request.POST.get('drug_num')
        # 获取图片
        drug_img=request.FILES.get('drug_img')
        # 获取进价
        purchasing_price=request.POST.get('purchasing_price')
        # 获取售价
        selling_price=request.POST.get('selling_price')
        # 获取药品名称
        drug_name=request.POST.get('drug_name')
        # 获取药品类型
        drug_type=request.POST.get('drug_type')
        # 获取简述
        brief_description=request.POST.get('brief_description')
        # 获取过期日期
        expiration_time=request.POST.get('expiration_time')
        # 获取详细描述
        detail=request.POST.get('detail')
        # 获取生产厂商
        manufacturer=request.POST.get('manufacturer')
        # 获取服用说明
        taking_instructions=request.POST.get('taking_instructions')
        # 获取备注
        remark=request.POST.get('remark')
        # 判断药品编号是否唯一
        if Drug.objects.filter(drug_num=drug_num):
            return HttpResponse('此药品编号已存在')
        # 判断药品信息是否为空
        elif len(drug_num)<=0 or len(purchasing_price)<=0 or len(selling_price)<=0 or len(drug_name)<=0 or len(drug_type)<=0 or len(brief_description)<=0 or len(expiration_time)<=0 or len(detail)<=0 or len(manufacturer)<=0 or len(taking_instructions)<=0 or len(remark)<=0:
            return HttpResponse('所有信息都不能为空')
        else:
            # 保存数据
            Drug.objects.create(drug_num=drug_num,drug_img=drug_img,purchasing_price=purchasing_price,selling_price=selling_price,drug_name=drug_name,drug_type=drug_type,brief_description=brief_description,expiration_time=expiration_time,detail=detail,manufacturer=manufacturer,taking_instructions=taking_instructions,remark=remark)
            return redirect('medicines:index')
# 展示
class index(views.View):
    def get(self,request):
        # 获取Drug表中所有数据
        drug_all = Drug.objects.all()
        limit = 2
        paginator = Paginator(drug_all, limit)  # 按每页10条分页
        page = request.GET.get('page', '1')  # 默认跳转到第一页
        result = paginator.page(page)
        return render(request, 'medicine/index.html',{'drug_all': result})
    def post(self,request):
        # 模糊查询
        pname = request.POST.get('pname')
        p_type = request.POST.get('ptype')
        print(p_type)
        drug_all = Drug.objects.filter(drug_name__icontains=pname, drug_type__icontains=p_type)
        return render(request, 'medicine/index.html', {'drug_all': drug_all})



def dpro4(request):
    data = {}
    ids = request.POST.get('ids')
    id_list = []
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in ids:
        if i != ',':
            id_list.append(i)
        else:
            pass
    for i in id_list:
        hos = Drug.objects.filter(id=int(i))
        ws['A1'].value = '药品编号'
        ws['B1'].value = '药品名称'
        ws['C1'].value = '药品类型'
        ws['D1'].value = '简单描述'
        ws['E1'].value = '状态'
        ws['F1'].value = '剩余量'
        row = 2
        for ch in hos:
            ws['A' + str(row)].value = ch.drug_num
            ws['B' + str(row)].value = ch.drug_name
            ws['C' + str(row)].value = ch.drug_type
            ws['D' + str(row)].value = ch.brief_description
            ws['E' + str(row)].value = ch.state
            ws['F' + str(row)].value = ch.surplus
            row += 1
            wb.save(f'{ch.drug_name}.xlsx')
            f = open(f'{ch.drug_name}.xlsx', 'rb')
            p = HttpResponse()
            p.content = f
            p['Content-Type'] = 'application/octet-stream'
            p['Content-Disposition'] = 'attachment;filename="%s.xlsx"' % ch.drug_name
    data['ids'] = ids
    return JsonResponse(data)
# # 更改药品
class Update(views.View):
    def get(self,request,u_id):
        # 获取Drug表中的id字段
        u =Drug.objects.get(id=u_id)
        return render(request, 'medicine/update.html',{'u':u})
    def post(self,request,u_id):
        # 获取编号
        drug_num=request.POST.get('drug_num')
        # 获取图片
        drug_img=request.POST.get('drug_img')
        # 获取进价
        purchasing_price=request.POST.get('purchasing_price')
        # 获取售价
        selling_price=request.POST.get('selling_price')
        # 获取药品名称
        drug_name=request.POST.get('drug_name')
        # 获取药品类型
        drug_type=request.POST.get('drug_type')
        # 获取简述
        brief_description=request.POST.get('brief_description')
        # 获取过期日期
        expiration_time=request.POST.get('expiration_time')
        # 获取详细描述
        detail=request.POST.get('detail')
        # 获取生产厂商
        manufacturer=request.POST.get('manufacturer')
        # 获取服用说明
        taking_instructions=request.POST.get('taking_instructions')
        # 获取备注
        remark=request.POST.get('remark')
        # 获取id
        u=Drug.objects.get(id=u_id)
        # 判断药品编号是否唯一
        if Drug.objects.filter(drug_num=drug_num):
            return HttpResponse('此药品编号已存在')
        # 判断药品信息是否为空
        elif len(drug_num) <= 0 or len(purchasing_price) <= 0 or len(selling_price) <= 0 or len(drug_name) <= 0 or len(
                drug_type) <= 0 or len(brief_description) <= 0 or len(expiration_time) <= 0 or len(detail) <= 0 or len(
                manufacturer) <= 0 or len(taking_instructions) <= 0 or len(remark) <= 0:
            return HttpResponse('所有信息都不能为空')
        else:
            # 更新数据
            u.drug_num=drug_num
            u.drug_img=drug_img
            u.purchasing_price=purchasing_price
            u.selling_price=selling_price
            u.drug_name=drug_name
            u.drug_type=drug_type
            u.brief_description=brief_description
            u.expiration_time=expiration_time
            u.detail=detail
            u.manufacturer=manufacturer
            u.taking_instructions=taking_instructions
            u.remark=remark
            u.save()
            return redirect('medicines:index')

# 展示详情
class look(views.View):
    def get(self,request,u_id):
        # 获取Drug表中的id字段
        u = Drug.objects.get(id=u_id)
        # 获取Drug表中所有数据
        drug_all = Drug.objects.all()
        return render(request, 'medicine/look.html',{'drug_all':drug_all,'u':u})

# 添加库存
class add_medicine(views.View):
    def get(self,request,u_id):
        # 获取Drug表中的id字段
        u = Drug.objects.get(id=u_id)
        # 获取Drug表中所有数据
        drug_all = Drug.objects.all()
        return render(request, 'medicine/add_medicine.html',{'drug_all':drug_all,'u':u})
    def post(self,request,u_id):
        # 获取数量
        add=request.POST.get('add')
        # 获取Drug表中的id字段
        u = Drug.objects.get(id=u_id)
        # 判断剩余量是否超过库存
        if u.surplus < u.inventory or u.surplus>0:
            u.state='销售中'
            i = int(add) + u.surplus
            # 判断添加后数量是否超过库存
            if i <= u.inventory:
                u.surplus += int(add)
                u.save()
                return redirect('medicines:index')
            else:
                return HttpResponse('药品数量不能超过库存3000')
        else:
            return HttpResponse('药品数量已满，无需添加库存')

def delete(request, u_id):
    u = Drug.objects.get(id=u_id)
    u.delete()
    return redirect('medicines:index')