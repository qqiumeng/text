from django.shortcuts import render, redirect, HttpResponse, reverse
from django.http import JsonResponse
from medicine.models import *
import openpyxl
from django.core.paginator import Paginator
from .forms import HospitalForm
from hospital.models import *
from registration.models import Departments
from django.db.models import Q
import datetime

def index(request):
    return render(request, 'Users/login.html')


def index1(request):
    if request.method == 'GET':
        hospital = Hospital.objects.all()
        limit = 5
        pnginator = Paginator(hospital, limit)
        page = request.GET.get('page', 1)
        result = pnginator.page(page)
        return render(request, 'hospital/index1.html', {'hospital': result})
    else:
        medical_record_No = request.POST.get('medical_record_No')  # 病历号
        doctor_name = request.POST.get('doctor_name')  # 主治医生
        office = request.POST.get('office')  # 所属科室
        time1 = request.POST.get('time1')
        time2 = request.POST.get('time2')
        print(medical_record_No, doctor_name, office, time1, time2)
        if medical_record_No and doctor_name and office and time1 and time2 == '':
            hospital = Hospital.objects.all()
            limit = 5
            pnginator = Paginator(hospital, limit)
            page = request.GET.get('page', 1)
            result = pnginator.page(page)
            return render(request, 'hospital/index1.html', {'hospital': result})
        elif len(time1)>0 and len(time2)>0:
            year = time1[0:4]
            month = time1[5:7]
            day = time1[8:10]
            year2 = time2[0:4]
            month2 = time2[5:7]
            day2 = time2[8:10]
            g1 = datetime.date(int(year), int(month), int(day))
            g2 = datetime.date(int(year2), int(month2), int(day2))
            hospital = Hospital.objects.filter(into_data__range=(g1,g2))
            limit = 5
            pnginator = Paginator(hospital, limit)
            page = request.GET.get('page', 1)
            result = pnginator.page(page)
            return render(request, 'hospital/index1.html', {'hospital': result})
        else:
            hospital = Hospital.objects.filter(Q(id__in=medical_record_No) | Q(office_id__doctor_name=doctor_name) | Q(
                office_id__d_office__department_name=office))
            return render(request, 'hospital/index1.html', {'hospital': hospital})


def out_hospital(request):
    id_list = []
    ids = request.POST.get('ss')
    for i in ids:
        if i != ',':
            id_list.append(i)
        else:
            pass
    for i in id_list:
        Hospital.objects.get(case_id=int(i)).delete()

    return redirect('hospital:index1')


def dpro(request):
    id_list = []
    ids = request.POST.get('ss')
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in ids:
        if i != ',':
            id_list.append(i)
        else:
            pass
    for i in id_list:
        hos = Hospital.objects.filter(case_id=int(i))
        ws['A1'].value = '病历号'
        ws['B1'].value = '姓名'
        ws['C1'].value = '床位号'
        ws['D1'].value = '联系电话'
        ws['E1'].value = '主治医生'
        ws['F1'].value = '入院时间'
        ws['G1'].value = '科室'
        ws['H1'].value = '状态'
        row = 2
        for obj in hos:
            ws['A' + str(row)].value = obj.case_id.id
            ws['B' + str(row)].value = obj.case_id.name
            ws['C' + str(row)].value = obj.bed_num
            ws['D' + str(row)].value = obj.case_id.tel
            ws['E' + str(row)].value = obj.office_id.doctor_name
            ws['F' + str(row)].value = obj.into_data
            ws['G' + str(row)].value = obj.office_id.d_office.department_name
            ws['H' + str(row)].value = obj.state
            row += 1
            wb.save(f'{obj.case_id.name}.xlsx')
            f = open(f'{obj.case_id.name}.xlsx', 'rb')
            p = HttpResponse()
            p.content = f
            p['Content-Type'] = 'application/octet-stream'
            p['Content-Disposition'] = 'attachment;filename="%s.xlsx"' % obj.case_id.name
    return render(request, 'hospital/index1.html')


def dpro2(request):
    data = {}
    h_id = request.POST.get('object_id')
    wb = openpyxl.Workbook()
    ws = wb.active
    Char = Charge.objects.filter(case_id_id=h_id)
    cash = request.POST.get('cash')
    balance = request.POST.get('balance')
    money = request.POST.get('money')
    ws['A1'].value = '病历号'
    ws['B1'].value = '姓名'
    ws['C1'].value = '收费项目'
    ws['D1'].value = '收费金额'
    ws['E1'].value = '收费日期'
    ws['F1'].value = '总花费'
    ws['G1'].value = '押金'
    ws['H1'].value = '余额'
    row = 2
    for ch in Char:
        ws['A' + str(row)].value = ch.case_id
        ws['B' + str(row)].value = ch.case_id.name
        ws['C' + str(row)].value = ch.charge_cellectable
        ws['D' + str(row)].value = ch.charge_amount
        ws['E' + str(row)].value = ch.exam_date
        ws['F' + str(row)].value = money
        ws['G' + str(row)].value = cash
        ws['H' + str(row)].value = balance
        row += 1
        wb.save(f'{ch.case_id.name}.xlsx')
        f = open(f'{ch.case_id.name}.xlsx', 'rb')
        p = HttpResponse()
        p.content = f
        p['Content-Type'] = 'application/octet-stream'
        p['Content-Disposition'] = 'attachment;filename="%s.xlsx"' % ch.case_id.name
    data['h_id'] = h_id
    return JsonResponse(data)


def edit_status(request):
    h_id = request.POST.get('object_id')
    new_statu = Hospital.objects.get(case_id_id=h_id)
    new_statu.state = '已结算'
    new_statu.save()
    return redirect('hospital:account')


def dpro3(request):
    ids = request.POST.get('h_id')
    id_list = []
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in ids:
        if i != ',':
            id_list.append(i)
        else:
            pass
    for i in id_list:
        hos = Hospital.objects.filter(case_id=int(i))
        ws['A1'].value = '病历号'
        ws['B1'].value = '姓名'
        ws['C1'].value = '押金'
        ws['D1'].value = '当前余额'
        ws['E1'].value = '状态'
        row = 2
        for obj in hos:
            ws['A' + str(row)].value = obj.case_id
            ws['B' + str(row)].value = obj.case_id.name
            ws['C' + str(row)].value = obj.cash
            ws['D' + str(row)].value = obj.case_id.money
            ws['E' + str(row)].value = obj.state
            row += 1
            wb.save(f'{obj.case_id.name}.xlsx')
            f = open(f'{obj.case_id.name}.xlsx', 'rb')
            p = HttpResponse()
            p.content = f
            p['Content-Type'] = 'application/octet-stream'
            p['Content-Disposition'] = 'attachment;filename="%s.xlsx"' % obj.case_id.name
    return redirect('hospital:account')


def edit(request, h_id):
    if request.method == 'GET':
        d_id = Patient.objects.get(id=h_id).id
        hospital = Hospital.objects.get(case_id=d_id)
        departments = Departments.objects.all()
        return render(request, 'hospital/edit.html', {'hospital': hospital, 'departments': departments})
    else:
        print('新的h_id字啊则热', h_id)
        d_id = Patient.objects.get(id=h_id).id
        doctor = Doctor.objects.all()
        hospital = Hospital.objects.get(case_id_id=d_id)
        doctor_name = request.POST.get('n1')
        d_id = Doctor.objects.get(doctor_name=doctor_name).id
        nurse = request.POST.get('nurse')
        bed_num = request.POST.get('bed_num')
        remark = request.POST.get('beizhu')
        is_t = request.POST.get('pname1')
        illness = request.POST.get('bingqing')
        p_info = Patient.objects.get(id=h_id)
        p_info.visit = is_t
        p_info.remark = remark
        hospital.nurse = nurse
        hospital.bed_num = bed_num
        hospital.illness = illness
        hospital.office_id_id = d_id
        p_info.save()
        hospital.save()
        return render(request, 'hospital/edit.html', {'hospital': hospital, 'doctor': doctor})


def account(request):
    if request.method == 'GET':

        hospital = Hospital.objects.all()
        limit = 2
        paginator = Paginator(hospital, limit)
        page = request.GET.get('page', '1')
        result = paginator.page(page)
        return render(request, 'hospital/account.html', {'hospital': result})
    else:
        me_no = request.POST.get('pname1')
        name = request.POST.get('pname2')
        hospital = Hospital.objects.filter(case_id_id=me_no, case_id__name=name)
        return render(request, 'hospital/account.html', {'hospital': hospital})


def account_look(request, h_id):
    money = 0
    char = Charge.objects.filter(case_id=h_id)
    all_money = char.values_list('charge_amount', flat=True)
    for i in all_money:
        money += i

    hosp = Hospital.objects.get(case_id_id=h_id)
    print(hosp, char)
    cash = hosp.cash

    balance = float(cash) - money
    return render(request, 'hospital/account-look.html',
                  {"char": char, "money": money, "cash": cash, "balance": balance, 'h_id': h_id})


def show_info(request):
    me_no = request.POST.get('me_no')
    data = {}
    if me_no != None:
        data['status'] = 'SUCCESS'
        data['me_no'] = me_no
        p_info = Patient.objects.get(id=me_no)
        data['name'] = p_info.name
        data['certificate_type'] = p_info.certificate_type
        data['id_num'] = p_info.id_num
        data['social_security_num'] = p_info.social_security_num
        data['tel'] = p_info.tel
        data['self_pay'] = p_info.self_pay
        data['sex'] = p_info.sex
        data['age'] = p_info.age
        data['visit'] = p_info.visit
        data['office'] = p_info.doctor_id.d_office.department_name
        data['doctor'] = p_info.doctor_id.doctor_name
        data['remark'] = p_info.remark
        print('me_no=', me_no)
        return JsonResponse(data)
    else:
        data['status'] = 'ERROR'
        return JsonResponse(data)


def add(request):
    if request.method == 'POST':
        hospital_form = HospitalForm(request.POST)
        hospital = Hospital.objects.all()
        if hospital_form.is_valid():
            nurse = hospital_form.cleaned_data['nurse']
            medical_record_No = request.POST.get('me_no')
            bed_num = hospital_form.cleaned_data['bed_num']
            cash = hospital_form.cleaned_data['cash']
            illness = hospital_form.cleaned_data['illness']
            p_info = Patient.objects.get(id=medical_record_No)
            Hospital.objects.create(bed_num=bed_num, cash=cash, state=p_info.state, nurse=nurse, illness=illness,
                                    case_id_id=p_info.id, office_id_id=p_info.doctor_id.id)
            return render(request, 'hospital/index1.html', {'hospital': hospital})
    else:
        hospital_form = HospitalForm()
    return render(request, 'hospital/add.html', {'hospital_form': hospital_form})


def add_many(request, h_id):
    if request.method == 'GET':
        hospital = Hospital.objects.get(id=h_id)
        return render(request, 'hospital/add_many.html', {'hospital': hospital})
    else:
        hospital = Hospital.objects.get(id=h_id)
        money = request.POST.get('money')
        hospital.case_id.money += float(money)
        hospital.save()
        return render(request, 'hospital/add_many.html', {'hospital': hospital})


def charge(request):
    if request.method == 'GET':
        char = Charge_manage.objects.all()
        limit = 2
        paginator = Paginator(char, limit)
        page = request.GET.get('page', '1')
        result = paginator.page(page)
        return render(request, 'hospital/charge.html', {'char': result})

    else:
        pname = request.POST.get('pname')
        char = Charge_manage.objects.filter(charge_cellectable=pname)
        return render(request, 'hospital/charge.html', {'char':char})


def charge_2(request):
    if request.method == 'GET':
        hosps = Hospital.objects.all()
        limit = 2
        paginator = Paginator(hosps, limit)
        page = request.GET.get('page', '1')
        result = paginator.page(page)
        return render(request, 'hospital/charge2.html', {'hosps': result})
    else:
        pname1 = request.POST.get('pname1')
        pname2 = request.POST.get('pname2')
        if Hospital.objects.filter(case_id_id=pname1, case_id__name=pname2).exists():
            hosps = Hospital.objects.filter(case_id_id=pname1, case_id__name=pname2)
            limit = 2
            paginator = Paginator(hosps, limit)
            page = request.GET.get('page', '1')
            result = paginator.page(page)
            return render(request, 'hospital/charge2.html', {'hosps': result})


def account_look2(request, h_id):
    hosp = {}
    hosps = Hospital.objects.filter(id=h_id)
    limit = 2
    paginator = Paginator(hosps, limit)  # 按每页10条分页
    page = request.GET.get('page', '1')  # 默认跳转到第一页
    result = paginator.page(page)
    hosp['hosps'] = result
    return render(request, 'hospital/account-look2.html', hosp)


def charge_edit1(request, c_id):
    if request.method == 'GET':
        char = Charge_manage.objects.get(id=c_id)
        return render(request, 'hospital/charge-edit1.html', {'char': char})
    else:
        char = Charge_manage.objects.get(id=c_id)
        charge_cellectable = request.POST.get('pname1')
        charge_amount = request.POST.get('pname2')
        if Charge_manage.objects.filter(charge_cellectable=charge_cellectable).exists():
            return render(request, 'hospital/charge-edit1.html', {'erro':'此项目已存在'})
        char.charge_cellectable = charge_cellectable
        char.charge_amount = charge_amount
        char.save()
        return redirect('hospital:charge')


def delete(request, c_id):
    a=Charge_manage.objects.get(id=c_id)
    Charge.objects.get(charge_manage_id=a.id).delete()
    a.delete()
    return redirect('hospital:charge')


def charge_edit2(request):
    if request.method == 'GET':

        return render(request, 'hospital/charge-edit2.html')
    else:
        pname1 = request.POST.get('pname1')
        pname2 = request.POST.get('pname2')
        if Charge_manage.objects.filter(charge_cellectable=pname1).exists():
            return render(request, 'hospital/charge-edit2.html', {'erro':'此项目已存在'})
        Charge_manage.objects.create(charge_cellectable=pname1, charge_amount=pname2, all_charge=pname2)
        return redirect('hospital:charge')


def charge_new(request, h_id):
    if request.method == 'GET':
        hosps = {}
        hosp = Hospital.objects.get(case_id=h_id)
        hosps['hosp'] = hosp
        return render(request, 'hospital/charge-new.html', hosps)
    else:
        hosps = {}
        pname1 = request.POST.get('pname1')
        pname2 = request.POST.get('pname2')
        print(pname1)
        charge_manage_id=Charge_manage.objects.get(charge_cellectable=pname1).id
        Charge.objects.create(charge_manage_id=charge_manage_id, case_id_id=h_id)
        hosp = Hospital.objects.get(case_id=h_id)
        hosps['hosp'] = hosp
        return render(request, 'hospital/charge-new.html', hosps)


def dispensing(request):
    if request.method == "GET":
        drug = Drug_grant.objects.all()
        print(drug)
        limit = 2
        paginator = Paginator(drug, limit)
        page = request.GET.get('page', '1')
        result = paginator.page(page)
        return render(request, 'hospital/dispensing.html', {'drug': result})
    else:
        me_no = request.POST.get('pname1')
        drug = Drug_grant.objects.filter(case_id_id=me_no)
        return render(request, 'hospital/dispensing.html', {'drug': drug})


def dispensing_give(request, h_id):
    if request.method == 'GET':
        drug_no = Drug_grant.objects.filter(case_id_id=h_id).values_list('case_id', flat=True)[0]
        drug_name = Drug_grant.objects.filter(case_id_id=h_id).values_list('case_id__name', flat=True)[0]
        drug_all = Drug.objects.all()

        return render(request, 'hospital/dispensing-give.html',
                      {'drug_no': drug_no, 'drug_all': drug_all, 'drug_name': drug_name})
    else:
        drug_all = Drug.objects.all()
        drug = Drug_grant.objects.filter(case_id_id=h_id)[0]
        drug_num = request.POST.get('drug_num')
        drug_name = request.POST.get('drug_name')
        drug1 = Drug.objects.get(drug_name=drug_name)
        if drug.not_sent > int(drug_num):
            drug1.surplus -= int(drug_num)
            drug.not_sent -= int(drug_num)
            drug.sent += int(drug_num)
        drug1.save()
        drug.save()
        return render(request, 'hospital/dispensing-give.html', {'drug': drug, 'drug_all': drug_all})


def dispend_all(request):
    data = {}
    d_id = request.POST.get('object_id')
    drug = Drug_grant.objects.get(id=d_id)
    drug.sent += drug.not_sent
    drug.not_sent -= drug.not_sent
    drug.save()
    return JsonResponse(data)


def dispensing_gives(request):
    return render(request, 'hospital/dispensing-gives.html')


def dispensing_look(request, h_id):
    drug = Drug_grant.objects.filter(case_id_id=h_id)

    return render(request, 'hospital/dispensing-look.html', {'drug': drug})


def look(request, h_id):
    hospital = Hospital.objects.get(id=h_id)
    return render(request, 'hospital/look.html', {'hospital': hospital})


def add_dis(request):
    if request.method == 'GET':
        return render(request, 'hospital/add_dispensing.html')
    else:
        num = request.POST.get('number')  # 病历号
        dis_name = request.POST.get('dis_name')  # 药品名称
        math = request.POST.get('drug_num')  # 发放总数
        mana = request.POST.get('manager')  # 负责人
        give_num = request.POST.get('give_num')  # 已发数量
        not_num = int(math)-int(give_num)
        try:
            case_id = Patient.objects.get(id=num).id
            if Drug.objects.filter(drug_name=dis_name).exists():
                drug_name_id = Drug.objects.get(drug_name=dis_name).id
                if int(give_num)>int(math):
                    return render(request, 'hospital/add_dispensing.html', {'erro': '已发数量大于总发放数，请重新输入'})
                else:
                    Drug_grant.objects.create(case_id_id=case_id, drug_name_id_id=drug_name_id, drup_num=math, sent=give_num,not_sent=not_num, principal=mana)
                return redirect('hospital:dispensing')
            else:
                return render(request, 'hospital/add_dispensing.html', {'erro': '此药品不存在，请重新输入'})
        except:
            return render(request, 'hospital/add_dispensing.html', {'erro': '此病号不存在，请重新输入'})





